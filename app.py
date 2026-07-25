from flask import Flask, render_template, request, redirect, url_for, session, flash, abort, Response
from supabase import create_client, Client
import qrcode
import io
import os
import base64
import threading
import time
from datetime import datetime, timezone
from werkzeug.utils import secure_filename
from dotenv import load_dotenv

# Set to "1" automatically by Vercel at runtime. Used below to adjust
# behaviour that only makes sense on a normal always-on server (background
# threads) vs. a serverless deployment (Vercel).
ON_VERCEL = bool(os.environ.get("VERCEL"))

# Load .env file FIRST
load_dotenv()

# Then read variables
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY")

print("URL:", SUPABASE_URL)
print("KEY:", SUPABASE_SERVICE_KEY[:20] + "...")  # Don't print the full key

supabase: Client = create_client(
    SUPABASE_URL,
    SUPABASE_SERVICE_KEY
)

app = Flask(__name__, template_folder='template')
# Reads FLASK_SECRET_KEY if set (recommended in production / Vercel env vars),
# otherwise falls back to the original hardcoded value so nothing breaks.
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "av_devlabs_secret")
app.config['MAX_CONTENT_LENGTH'] = 15 * 1024 * 1024  # 15 MB max upload size

# ---------------------------------------------------------------------------
# Cache-busting for /static/*.
# vercel.json sets "Cache-Control: public, max-age=31536000, immutable" on
# everything under /static/ for performance — but that means once a browser
# has fetched e.g. theme.css, it will NOT re-fetch it for a year, even after
# we deploy changes (not even on a hard refresh, in some browsers). To make
# updates actually show up, every static asset URL gets a ?v=<version> query
# string appended, computed from that file's last-modified time. Changing the
# file changes the URL, so it's always a fresh cache miss — the 1-year cache
# is preserved for assets that haven't changed, and busted for ones that have.
# ---------------------------------------------------------------------------
_STATIC_DIR = os.path.join(app.root_path, 'static')

def _static_version(filename):
    try:
        return str(int(os.path.getmtime(os.path.join(_STATIC_DIR, filename))))
    except OSError:
        return "1"

@app.context_processor
def inject_asset_version():
    def versioned_static(filename):
        return url_for('static', filename=filename) + '?v=' + _static_version(filename)
    return dict(versioned_static=versioned_static)


# Compress every HTML/CSS/JS/JSON response with gzip (falls back automatically
# if the browser doesn't support it). This cuts payload size dramatically for
# almost no CPU cost and helps a lot under concurrent load.
try:
    from flask_compress import Compress
    Compress(app)
except ImportError:
    pass

# Tell browsers to cache static files (images, css) for a year. Because the
# upload routes already give files random uuid-based filenames, a changed
# file gets a new URL automatically, so long caching here is safe and means
# repeat visitors barely re-download anything.
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 60 * 60 * 24 * 365

BASE_DIR = os.path.abspath(os.path.dirname(__file__))

# ---------------------------------------------------------------------------
# Supabase Connection
# ---------------------------------------------------------------------------
# Set these two as environment variables (or in a local .env file):
#   SUPABASE_URL          -> Project Settings > API > Project URL
#   SUPABASE_SERVICE_KEY  -> Project Settings > API > service_role secret key
# The service_role key is used because this is a trusted server-side backend;
# never expose it in frontend/browser code.

TABLE_QUERIES = "queries"
TABLE_ORDERS = "place_orders"
TABLE_TRASH = "trash"
TABLE_ANNOUNCEMENTS = "announcements"
TABLE_REVIEWS = "reviews"
TABLE_REPORT_LEADS = "report_leads"

# Contact number for the company (used for the "Chat on WhatsApp" button)
COMPANY_WHATSAPP_NUMBER = "919825089454"

# Service categories shown on the "Place Order" form. Visitors can also type
# a custom category via "Other" if theirs isn't listed.
SERVICE_CHOICES = [
    "Website Making",
    "App Making",
    "ERP Software",
    "AI Agent",
]


# ---------------------------------------------------------------------------
# Tiny TTL cache for hot, read-heavy pages (homepage announcements, shop
# products). These are read on nearly every visitor request but only ever
# change when the admin edits them, so caching for a few seconds massively
# cuts the number of Supabase calls when many people browse at once, without
# ever showing data more than a few seconds stale.
# ---------------------------------------------------------------------------
_cache_store = {}
_cache_lock = threading.Lock()


def cached(key, ttl_seconds, fetch_fn):
    now = time.time()
    with _cache_lock:
        entry = _cache_store.get(key)
        if entry and entry[0] > now:
            return entry[1]
    value = fetch_fn()
    with _cache_lock:
        _cache_store[key] = (now + ttl_seconds, value)
    return value


def invalidate_cache(key):
    with _cache_lock:
        _cache_store.pop(key, None)


def _with_alias(rows):
    """Supabase rows use 'id' (bigint). Existing templates were written against
    MongoDB's '_id' field, so we mirror it as a string here to keep every
    template working unchanged."""
    for r in rows or []:
        if r and 'id' in r:
            r['_id'] = str(r['id'])
    return rows or []


def _one_or_none(rows):
    return rows[0] if rows else None


def _strip_row_id(row):
    """Remove identity/meta columns before re-inserting a row into another table."""
    if not row:
        return row
    row = dict(row)
    row.pop('id', None)
    row.pop('_id', None)
    row.pop('created_at', None)
    return row


# ---------------------------------------------------------------------------
# REPORT GENERATION PIPELINE
# ---------------------------------------------------------------------------
# Every time the Query form or the Place Order form is submitted, we ALSO
# copy just the fields needed for reporting (name, gmail, category, source)
# into a separate `report_leads` table. This is a one-way, additive side
# copy — it never reads from or modifies `queries` / `place_orders`, so the
# original submit workflows and data are completely unaffected. Report
# generation (Admin > Reports) only ever reads from `report_leads`.
#
# De-duplication by Gmail: `report_leads.gmail` has a UNIQUE constraint in
# the database, so this uses an upsert — if a row with the same email
# already exists, it is UPDATED in place with the latest name/category/
# source. A new/different email always creates a fresh, unique row. This
# keeps the report table at exactly one row per real lead, no matter how
# many forms they've filled.
def sync_lead_report(name, gmail, category, source):
    if not gmail:
        # Without an email there's nothing to de-duplicate on, so skip the
        # report copy rather than risk creating junk/duplicate rows.
        return
    try:
        supabase.table(TABLE_REPORT_LEADS).upsert({
            "name": name,
            "gmail": gmail,
            "category": category,
            "source": source,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }, on_conflict="gmail").execute()
    except Exception as e:
        # The report pipeline must NEVER break the real query/order workflow,
        # so any failure here is only logged, never raised.
        print("[report pipeline] could not sync report_leads row:", e)


def _resolve_category(form):
    """Read category_choice (+ optional new_category for 'Other') from the
    Place Order form, the same '+ Create New' pattern used for product
    categories previously."""
    category_choice = (form.get('category_choice') or '').strip()
    new_category = (form.get('new_category') or '').strip()
    if category_choice == '__other__':
        return new_category
    return category_choice


def count_rows(table, **filters):
    q = supabase.table(table).select("id", count="exact")
    for key, value in filters.items():
        q = q.eq(key, value)
    return q.execute().count or 0


def get_admin_counts():
    # These 5 counts are independent, so run them concurrently on a small
    # thread pool instead of one-after-another - cuts this from ~5 sequential
    # network round-trips to ~1, which matters a lot when several admins/
    # pages are loading at once.
    from concurrent.futures import ThreadPoolExecutor

    jobs = {
        "orders_pending": lambda: count_rows(TABLE_ORDERS, status="Pending"),
        "trash": lambda: count_rows(TABLE_TRASH),
        "announcements": lambda: count_rows(TABLE_ANNOUNCEMENTS),
        "reviews_pending": lambda: count_rows(TABLE_REVIEWS, status="Pending"),
    }
    with ThreadPoolExecutor(max_workers=4) as pool:
        futures = {key: pool.submit(fn) for key, fn in jobs.items()}
        return {key: f.result() for key, f in futures.items()}


def login_required(view_func):
    from functools import wraps

    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return view_func(*args, **kwargs)
    return wrapped


# ---------------------------------------------------------------------------
# HOME PAGE + QUERY SYSTEM
# ---------------------------------------------------------------------------
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        name = request.form.get('name')
        gmail = (request.form.get('gmail') or '').strip()
        query_text = request.form.get('query')

        new_query = {
            "name": name,
            "gmail": gmail,
            "query": query_text,
            "status": "Pending",
            "admin_message": "",
            "date": datetime.now().strftime("%d %b %Y, %I:%M %p")
        }
        supabase.table(TABLE_QUERIES).insert(new_query).execute()

        # Report generation pipeline: copy just the report-relevant fields
        # into report_leads (deduped by Gmail). The original query row/
        # workflow above is already saved and unaffected by this.
        sync_lead_report(name, gmail, "General Query", "Query Form")

        flash("Query submitted successfully! 🚀", "success")
        return redirect(url_for('index'))

    def _fetch():
        resp = supabase.table(TABLE_ANNOUNCEMENTS).select("*").order("id", desc=True).limit(12).execute()
        return _with_alias(resp.data)

    announcements = cached("home_announcements", 30, _fetch)

    def _fetch_reviews():
        resp = supabase.table(TABLE_REVIEWS).select("*").eq("status", "Approved").order("id", desc=True).limit(12).execute()
        return _with_alias(resp.data)

    reviews = cached("home_reviews", 30, _fetch_reviews)

    return render_template('index.html', whatsapp_number=COMPANY_WHATSAPP_NUMBER, announcements=announcements,
                            reviews=reviews, service_choices=SERVICE_CHOICES)


# ---------------------------------------------------------------------------
# PLACE ORDER (project request form)
# ---------------------------------------------------------------------------
@app.route('/place-order', methods=['POST'])
def place_order():
    name = request.form.get('name')
    contact_number = request.form.get('contact_number')
    gmail = (request.form.get('gmail') or '').strip()
    special_instructions = request.form.get('special_instructions')
    category = _resolve_category(request.form)

    clean_number = ''.join(filter(str.isdigit, contact_number or ''))

    supabase.table(TABLE_ORDERS).insert({
        "name": name,
        "contact_number": clean_number,
        "gmail": gmail,
        "category": category,
        "special_instructions": special_instructions,
        "status": "Pending",
        "date": datetime.now().strftime("%d %b %Y, %I:%M %p")
    }).execute()

    # Report generation pipeline: copy just the report-relevant fields into
    # report_leads (deduped by Gmail). The original order row/workflow above
    # is already saved and unaffected by this.
    sync_lead_report(name, gmail, category or "Other", "Place Order")

    flash("Your project request has been submitted! We'll get back to you soon. 🚀", "success")
    return redirect(url_for('index') + '#placeOrderForm')


# ---------------------------------------------------------------------------
# LOGIN SYSTEM (admin only)
# ---------------------------------------------------------------------------
# Deliberately NOT linked from any visible nav/footer button, per the "hide
# the admin panel" requirement — admins reach it by typing this URL directly.
# The route name stays 'login' (so url_for('login') keeps working anywhere
# it's referenced), only the URL path itself is the hidden part.
@app.route('/av-admin-access', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if username == 'admin' and password == 'admin123':
            session['logged_in'] = True
            return redirect(url_for('admin'))
        else:
            flash("Invalid Credentials. Please try again.", "danger")

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))


# ---------------------------------------------------------------------------
# ADMIN - QUERY DASHBOARD
# ---------------------------------------------------------------------------
@app.route('/admin')
@login_required
def admin():
    search_q = request.args.get('q', '').strip()

    q = supabase.table(TABLE_QUERIES).select("*").order("id", desc=True)
    if search_q:
        q = q.ilike("name", f"%{search_q}%")
    all_queries = _with_alias(q.execute().data)

    stats = {
        "total": count_rows(TABLE_QUERIES),
        "pending": count_rows(TABLE_QUERIES, status="Pending"),
        "available": count_rows(TABLE_QUERIES, status="Available"),
        "non_available": count_rows(TABLE_QUERIES, status="Non-Available"),
    }

    return render_template('admin.html', queries=all_queries, stats=stats, counts=get_admin_counts(), search_q=search_q)


@app.route('/update_status/<query_id>', methods=['POST'])
@login_required
def update_status(query_id):
    new_status = request.form.get('status')
    admin_message = request.form.get('admin_message')

    supabase.table(TABLE_QUERIES).update({
        "status": new_status,
        "admin_message": admin_message
    }).eq("id", int(query_id)).execute()

    flash("Query updated successfully! ✅", "success")
    return redirect(url_for('admin'))

@app.route('/delete_query/<int:query_id>', methods=['POST'])
@login_required
def delete_query(query_id):
    try:
        supabase.table(TABLE_QUERIES).delete().eq("id", query_id).execute()
        flash("Query deleted successfully. 🗑️", "success")
    except Exception as e:
        flash(f"Error deleting query: {e}", "danger")

    return redirect(url_for('admin'))

# ---------------------------------------------------------------------------
# ANNOUNCEMENTS (circular cards shown on the homepage)
# ---------------------------------------------------------------------------
ANNOUNCEMENT_COLORS = {'amber', 'green', 'blue', 'red', 'purple'}
ANNOUNCEMENT_ICONS = [
    'bi-megaphone-fill', 'bi-gift-fill', 'bi-percent', 'bi-stars',
    'bi-calendar-event-fill', 'bi-bell-fill', 'bi-tag-fill', 'bi-emoji-smile-fill'
]


@app.route('/admin/announcements')
@login_required
def admin_announcements():
    resp = supabase.table(TABLE_ANNOUNCEMENTS).select("*").order("id", desc=True).execute()
    all_announcements = _with_alias(resp.data)
    return render_template('admin_announcements.html', announcements=all_announcements,
                            counts=get_admin_counts(), icon_choices=ANNOUNCEMENT_ICONS)


@app.route('/admin/announcements/add', methods=['POST'])
@login_required
def admin_announcements_add():
    title = request.form.get('title', '').strip()
    message = request.form.get('message', '').strip()
    icon = request.form.get('icon', '').strip() or 'bi-megaphone-fill'
    color = request.form.get('color', '').strip()
    if color not in ANNOUNCEMENT_COLORS:
        color = 'amber'

    if not title:
        flash("Announcement title is required.", "danger")
        return redirect(url_for('admin_announcements'))

    supabase.table(TABLE_ANNOUNCEMENTS).insert({
        "title": title,
        "message": message,
        "icon": icon,
        "color": color,
        "date": datetime.now().strftime("%d %b %Y, %I:%M %p")
    }).execute()
    invalidate_cache("home_announcements")
    flash("Announcement posted! 📢", "success")
    return redirect(url_for('admin_announcements'))


@app.route('/admin/announcements/delete/<announcement_id>', methods=['POST'])
@login_required
def admin_announcements_delete(announcement_id):
    supabase.table(TABLE_ANNOUNCEMENTS).delete().eq("id", int(announcement_id)).execute()
    invalidate_cache("home_announcements")
    flash("Announcement removed.", "success")
    return redirect(url_for('admin_announcements'))


# ---------------------------------------------------------------------------
# REVIEWS (customer feedback -> curated "Our Happy Customers" homepage section)
# ---------------------------------------------------------------------------
@app.route('/review', methods=['GET', 'POST'])
def review():
    if request.method == 'POST':
        name = request.form.get('name')
        whatsapp_number = request.form.get('whatsapp_number')
        review_text = request.form.get('review_text')
        try:
            rating = max(1, min(5, int(request.form.get('rating', 5))))
        except (TypeError, ValueError):
            rating = 5

        clean_number = ''.join(filter(str.isdigit, whatsapp_number))

        supabase.table(TABLE_REVIEWS).insert({
            "name": name,
            "whatsapp_number": clean_number,
            "review_text": review_text,
            "rating": rating,
            "status": "Pending",
            "date": datetime.now().strftime("%d %b %Y, %I:%M %p")
        }).execute()

        flash("Thank you for your feedback! Your review has been submitted. 🌟", "success")
        return redirect(url_for('review'))

    return render_template('review.html', whatsapp_number=COMPANY_WHATSAPP_NUMBER)


@app.route('/admin/reviews')
@login_required
def admin_reviews():
    status_filter = request.args.get('status', '').strip()
    q = supabase.table(TABLE_REVIEWS).select("*").order("id", desc=True)
    if status_filter:
        q = q.eq("status", status_filter)
    all_reviews = _with_alias(q.execute().data)

    stats = {
        "total": count_rows(TABLE_REVIEWS),
        "pending": count_rows(TABLE_REVIEWS, status="Pending"),
        "approved": count_rows(TABLE_REVIEWS, status="Approved"),
        "deleted": count_rows(TABLE_REVIEWS, status="Deleted"),
    }
    return render_template('admin_reviews.html', reviews=all_reviews, stats=stats,
                            counts=get_admin_counts(), status_filter=status_filter)


@app.route('/admin/reviews/update/<review_id>', methods=['POST'])
@login_required
def admin_reviews_update(review_id):
    new_status = request.form.get('status')
    supabase.table(TABLE_REVIEWS).update({"status": new_status}).eq("id", int(review_id)).execute()
    invalidate_cache("home_reviews")
    if new_status == "Approved":
        flash("Review approved and posted to Our Happy Customers! ⭐", "success")
    elif new_status == "Deleted":
        flash("Review marked as deleted.", "success")
    else:
        flash("Review set to pending.", "success")
    return redirect(url_for('admin_reviews'))


@app.route('/admin/reviews/delete/<review_id>', methods=['POST'])
@login_required
def admin_reviews_delete(review_id):
    supabase.table(TABLE_REVIEWS).delete().eq("id", int(review_id)).execute()
    invalidate_cache("home_reviews")
    flash("Review permanently removed. 🗑️", "success")
    return redirect(url_for('admin_reviews'))



# ---- Admin: Orders (submitted carts) ----
@app.route('/admin/orders')
@login_required
def admin_orders():
    resp = supabase.table(TABLE_ORDERS).select("*").order("id", desc=True).execute()
    all_orders = _with_alias(resp.data)
    return render_template('admin_orders.html', orders=all_orders, counts=get_admin_counts())


@app.route('/admin/orders/done/<order_id>', methods=['POST'])
@login_required
def admin_orders_done(order_id):
    supabase.table(TABLE_ORDERS).update({"status": "Done"}).eq("id", int(order_id)).execute()
    flash("Order marked as done. ✅", "success")
    return redirect(url_for('admin_orders'))


@app.route('/admin/orders/remove/<order_id>', methods=['POST'])
@login_required
def admin_orders_remove(order_id):
    resp = supabase.table(TABLE_ORDERS).select("*").eq("id", int(order_id)).execute()
    order = _one_or_none(resp.data)
    if order:
        trash_row = _strip_row_id(order)
        trash_row['original_order_id'] = order['id']
        trash_row['deleted_date'] = datetime.now().strftime("%d %b %Y, %I:%M %p")
        supabase.table(TABLE_TRASH).insert(trash_row).execute()
        supabase.table(TABLE_ORDERS).delete().eq("id", int(order_id)).execute()
        flash("Order moved to Trash.", "success")
    return redirect(url_for('admin_orders'))


# ---- Admin: Trash ----
@app.route('/admin/trash')
@login_required
def admin_trash():
    resp = supabase.table(TABLE_TRASH).select("*").order("id", desc=True).execute()
    trashed = _with_alias(resp.data)
    return render_template('admin_trash.html', trashed=trashed, counts=get_admin_counts())


@app.route('/admin/trash/restore/<order_id>', methods=['POST'])
@login_required
def admin_trash_restore(order_id):
    resp = supabase.table(TABLE_TRASH).select("*").eq("id", int(order_id)).execute()
    order = _one_or_none(resp.data)
    if order:
        restored_row = _strip_row_id(order)
        restored_row.pop('deleted_date', None)
        restored_row.pop('original_order_id', None)
        restored_row['status'] = 'Pending'
        supabase.table(TABLE_ORDERS).insert(restored_row).execute()
        supabase.table(TABLE_TRASH).delete().eq("id", int(order_id)).execute()
        flash("Order restored.", "success")
    return redirect(url_for('admin_trash'))


@app.route('/admin/trash/clear', methods=['POST'])
@login_required
def admin_trash_clear():
    supabase.table(TABLE_TRASH).delete().gt("id", 0).execute()
    flash("Trash bin cleared.", "success")
    return redirect(url_for('admin_trash'))


# ---------------------------------------------------------------------------
# ADMIN - REPORT GENERATION (Excel)
# ---------------------------------------------------------------------------
# Reads ONLY from report_leads (never from queries / place_orders), so
# this feature can never touch or exploit the original request data/workflow.
@app.route('/admin/reports')
@login_required
def admin_reports():
    resp = supabase.table(TABLE_REPORT_LEADS).select("*").execute()
    rows = resp.data or []

    categories = sorted({r['category'] for r in rows if r.get('category')})
    sources = sorted({r['source'] for r in rows if r.get('source')})

    return render_template(
        'admin_reports.html',
        counts=get_admin_counts(),
        total_leads=len(rows),
        categories=categories,
        sources=sources,
    )


@app.route('/admin/reports/export')
@login_required
def admin_reports_export():
    filter_type = request.args.get('filter_type', 'all')
    filter_value = request.args.get('filter_value', '').strip()

    q = supabase.table(TABLE_REPORT_LEADS).select("*")
    sheet_title = "All Leads"
    if filter_type == 'category' and filter_value:
        q = q.eq('category', filter_value)
        sheet_title = f"Category - {filter_value}"
    elif filter_type == 'source' and filter_value:
        q = q.eq('source', filter_value)
        sheet_title = f"Source - {filter_value}"

    rows = q.order('name').execute().data or []

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Lead Report"  # Excel sheet-name limit

    headers = ["Sr No.", "Name", "Gmail", "Category", "Source"]
    ws.append(headers)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows, start=1):
        ws.append([
            i,
            r.get('name') or '',
            r.get('gmail') or '',
            r.get('category') or '',
            r.get('source') or '',
        ])

    for idx, width in enumerate([8, 26, 30, 24, 16], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename_bits = ["lead_report"]
    if filter_type != 'all' and filter_value:
        filename_bits.append(filter_type)
        filename_bits.append(secure_filename(filter_value).lower() or "value")
    filename_bits.append(datetime.now().strftime("%Y%m%d_%H%M"))
    filename = "_".join(filename_bits) + ".xlsx"

    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ---------------------------------------------------------------------------
# QR CODE
# ---------------------------------------------------------------------------
@app.route('/generate_qr')
@login_required
def generate_qr():
    # Was hardcoded to http://127.0.0.1:5000, which only worked while running
    # locally. Using the current request's own domain means the generated QR
    # code correctly points at wherever the site is actually deployed
    # (Vercel URL, custom domain, or localhost during local development).
    website_url = request.url_root
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(website_url)
    qr.make(fit=True)
    img = qr.make_image(fill='black', back_color='white')

    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")

    return render_template('qr_code.html', qr_image=qr_base64)


if __name__ == '__main__':
    app.run(debug=True)
